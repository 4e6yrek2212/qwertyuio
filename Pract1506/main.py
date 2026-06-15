import requests
import matplotlib.pyplot as plt
import datetime
import openpyxl
import json
import re
import random
import time
from colorama import Fore, Style
from collections import Counter
from openpyxl.styles import Alignment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

HH_API_URL = "https://api.hh.ru/vacancies"

def generate_mock_vacancies(keyword, count=20):
    print(f"⚠️ API заблокировано (403). Включен режим генерации тестовых данных для '{keyword}'...")

    cities = ['Москва', 'Санкт-Петербург', 'Новосибирск', 'Екатеринбург', 'Казань', 'Краснодар']
    employers = ['ТехноРешения', 'Диджитал Лаб', 'СберТех-Имитация', 'Яндекс.Старт', 'ФинТех Групп', 'Айти Вектор',
                 'Стартап 3000']

    skills_pool = ['Python', 'SQL', 'Git', 'Docker', 'Linux', 'REST API', 'PostgreSQL', 'Django', 'FastAPI']
    if 'frontend' in keyword.lower() or 'фронтенд' in keyword.lower() or 'javascript' in keyword.lower():
        skills_pool = ['JavaScript', 'TypeScript', 'React', 'Vue.js', 'HTML5', 'CSS3', 'Git', 'Webpack']
    elif 'qa' in keyword.lower() or 'тестировщик' in keyword.lower():
        skills_pool = ['QA', 'Selenium', 'Postman', 'Jira', 'Python', 'SQL', 'Allure', 'Git', 'Charles']
    elif 'data' in keyword.lower() or 'аналитик' in keyword.lower():
        skills_pool = ['Python', 'SQL', 'Pandas', 'NumPy', 'Tableau', 'PowerBI', 'A/B testing', 'Airflow']

    prefixes = ['Senior', 'Junior', 'Middle', 'Ведущий', 'Стажер']

    mock_data = []
    current_date = datetime.datetime.now().strftime("%d-%m-%Y")

    for i in range(count):
        pref = random.choice(prefixes)
        name = f"{pref} {keyword}" if random.random() > 0.3 else f"{keyword}"

        salary_from = random.randint(40, 220) * 1000
        salary_to = salary_from + random.randint(20, 80) * 1000 if random.random() > 0.4 else None

        selected_skills = random.sample(skills_pool, k=random.randint(2, 4))
        requirement = f"Опыт разработки от 2 лет. Требуются знания: {', '.join(selected_skills)}. Опыт работы с базами данных."

        vacancy_info = {
            'name': name,
            'salary': {
                'from': salary_from,
                'to': salary_to,
                'currency': 'RUB'
            },
            'city': random.choice(cities),
            'employer': random.choice(employers),
            'published_at': current_date,
            'requirement': requirement
        }
        mock_data.append(vacancy_info)

    return mock_data

def fetch_vacancies(keyword, pages=2):
    vacancies = []
    headers = {
        'User-Agent': 'VacancyAnalyzerApp/1.0 (test_user@example.com)'
    }

    for page in range(pages):
        params = {
            'text': keyword,
            'page': page,
            'per_page': 20,
            'area': 1
        }
        try:
            response = requests.get(HH_API_URL, params=params, headers=headers, timeout=5)

            if response.status_code == 403:
                # Если поймали 403 на любой странице, сразу переключаемся на генератор
                return generate_mock_vacancies(keyword)

            if response.status_code != 200:
                print(f"Ошибка API: {response.status_code}")
                continue

            data = response.json()
        except Exception as e:
            print(f"Ошибка соединения: {e}. Переключаюсь на генерацию...")
            return generate_mock_vacancies(keyword)

        for item in data.get('items', []):
            area_name = item.get('area', {}).get('name', 'Не указан')
            employer_name = item.get('employer', {}).get('name', 'Не указан')
            published_at = item.get('published_at', '')

            try:
                date_obj = datetime.datetime.strptime(published_at.split('T')[0], "%Y-%m-%d")
                published_formatted = date_obj.strftime("%d-%m-%Y")
            except ValueError:
                published_formatted = published_at

            requirement = item.get('snippet', {}).get('requirement', '')
            if requirement:
                requirement = re.sub(r'<[^>]+>', '', requirement)

            vacancy_info = {
                'name': item.get('name', 'Без названия'),
                'salary': item.get('salary'),
                'city': area_name,
                'employer': employer_name,
                'published_at': published_formatted,
                'requirement': requirement
            }
            vacancies.append(vacancy_info)

        time.sleep(1)  # Небольшая пауза между страницами

    return vacancies

def filter_by_salary(vacancies, min_salary=0):
    filtered = []
    for vac in vacancies:
        salary = vac.get('salary')
        if not salary:
            continue

        salary_from = salary.get('from')
        salary_to = salary.get('to')
        current_salary = salary_from if salary_from is not None else salary_to

        if current_salary and current_salary >= min_salary:
            filtered.append(vac)
    return filtered

def save_to_exel(vacancies, keyword):
    filename = f'vacancies_{keyword}.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = "Вакансии"

    # Теперь колонка 'Зарплата' одна
    header = ['Название', 'Зарплата', 'Город', 'Работодатель', 'Дата публикации', 'Требования']
    ws.append(header)

    for vacancy in vacancies:
        salary = vacancy.get('salary')

        # Форматируем зарплату в одну строку через дефис
        if salary:
            s_from = salary.get('from')
            s_to = salary.get('to')
            currency = salary.get('currency', 'RUB')

            if s_from and s_to:
                salary_text = f"{s_from} - {s_to} {currency}"
            elif s_from:
                salary_text = f"от {s_from} {currency}"
            elif s_to:
                salary_text = f"до {s_to} {currency}"
            else:
                salary_text = "Не указана"
        else:
            salary_text = "Не указана"

        row = [
            vacancy.get('name', ''),
            salary_text,  # Красивая единая строка зарплаты
            vacancy.get('city', ''),
            vacancy.get('employer', ''),
            vacancy.get('published_at', ''),
            vacancy.get('requirement', '')
        ]
        ws.append(row)

    # Автоматическая настройка ширины и перенос текста
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            # Для колонки "Требования" (она у нас 6-я по счету, то есть 'F') делаем перенос текста
            if col_letter == 'F' and cell.row > 1:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                continue  # Не учитываем длину текста требований в общей авто-ширине

            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        # Задаем ширину: для требований фиксируем 40 (с переносом), для остальных — авторазмер
        if col_letter == 'F':
            ws.column_dimensions[col_letter].width = 40
        else:
            ws.column_dimensions[col_letter].width = min(max_length + 3, 50)

    wb.save(filename)
    print(f'📄 Результаты сохранены в файл: {filename}')

def get_top_skills(vacancies, top_n=5):
    words = []
    for vac in vacancies:
        requirement = vac.get('requirement') or ''
        # Ищем технологии на латинице (включая C++, C#, .NET)
        tech_words = re.findall(r'\b[a-zA-Z]+(?:[+#]{1,2}|\.net)?\b', requirement, re.IGNORECASE)
        words.extend([word.upper() for word in tech_words])

    counter = Counter(words)
    return counter.most_common(top_n)

def get_graph(data, keyword):
    names = []
    salaries = []

    for vac in data:
        salary = vac.get('salary')
        if salary and salary.get('from'):
            label = f"{vac.get('name')} \n({vac.get('employer')})"
            names.append(label)
            salaries.append(salary.get('from'))

    names = names[:8]
    salaries = salaries[:8]

    plt.figure(figsize=(8, 4))
    plt.bar(names, salaries, color='#4A90E2')
    plt.title(f"Зарплаты по запросу: {keyword}", fontsize=12, fontweight='bold')
    plt.xlabel("Рубли")
    plt.grid(axis='x', linestyle='--', alpha=0.6)
    plt.tight_layout()
    plt.show()

def get_statistics(vacancies):
    pass

def main():

    while True:
        print("\n1. Найти вакансии\n2. Самые популярные\n3. Сравнение МСК и СПБ\n4. График зарплат\n5. Выход")
        key = input("Введите номер команды: ")

        if key == '1':
            keyword = input("Введите профессию для поиска: ")
            print("Сбор вакансий...")
            data = fetch_vacancies(keyword, pages=3)
            print(f"Получено всего: {len(data)} вакансий")

            if not data:
                print("Нет данных для анализа.")
                return

            filtered_data = filter_by_salary(data, min_salary=50000)
            print(f"После фильтрации по ЗП (от 50 000): {len(filtered_data)}")

            if filtered_data:
                save_to_exel(filtered_data, keyword)
            else:
                print("Нет данных для сохранения в Excel после фильтрации.")

        elif key == '2':
            print("\n🔥 Топ-5 навыков из описания вакансий:")
            top_skills = get_top_skills(data, top_n=5)
            for idx, (skill, count) in enumerate(top_skills, 1):
                print(f"{idx}. {Fore.RED}{skill} {Style.RESET_ALL}— упоминается {count} раз(а)")

        elif key == '3':
            pass

        elif key == '4':
            keyword = input("Введите профессию: ")
            data = []
            data = fetch_vacancies(keyword)
            get_graph(data, keyword)

        elif key == '5':
            exit()

if __name__ == "__main__":
    main()
